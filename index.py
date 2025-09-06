# import random
import openpyxl as xl
# from pathlib import Path
# # import time
# #
# # def game():
# #   ran = random.randint(0, 10)
# #   attempts = 0
# #   attempts_limit = 5
# #
# #   while attempts < attempts_limit:
# #     user = int(input("Pick A Number: "))
# #     attempts += 1
# #
# #     if user == ran:
# #       print(f"Correct! You picked {user}")
# #       break
# #   else:
# #     print("Sorry you Failed!")
# #
# # # game()
# #
# # def car_game():
# #     user_input = ""
# #     started = False
# #     while True:
# #         user_input = input("> ").lower()
# #
# #         if user_input == "start":
# #             if started:
# #                 print("Car is already started!")
# #             else:
# #                 started = True
# #                 print("Car is started.")
# #         elif user_input == "stop":
# #             if not started:
# #                 print("Car is already stopped!")
# #             else:
# #                 started = False
# #                 print("Car is stopped.")
# #         elif user_input == "exit":
# #             print("The game has been exited")
# #             break
# #         elif user_input == "help":
# #             print("""
# # start - To start the car
# # stop - To stop the car
# # exit - To exit the game""")
# #         else:
# #             print("Please enter a valid input")
# #
# # # car_game()
# #
# # def main():
# #     item = [3, 4, 5, 3, 1, 4]
# #     total = 0
# #     for i in item:
# #         total += i
# #     print(f"The total is: {total}")
# #
# # # main()
# #
# # # for i in [5, 1, 3, 1, 1]:
# #   #  print(i * "x")
# #
# # def user_name():
# #     while True:
# #         name = input("What is your name: ").upper()
# #         if name == "":
# #             print('Please enter your name!')
# #         else:
# #             print(f"Your name is {name}")
# #             break
# #
# # # user_name()
# #
# #
# # def lift_off():
# #     time_limit = int(input("Pick a time limit: "))
# #     for numbers in range(time_limit):
# #         print(numbers)
# #         time_limit =- 1
# #         time.sleep(1)
# #     print("Lift Off!!!")
# #
# # # lift_off()
# #
# # class Person:
# #     def __init__(self, name):
# #         self.name = name
# #     def talk(self):
# #         print(f"I am {self.name}")
# #
# # ismeta = Person("Clinton")
# # ismeta.talk()
#
# # number = random.randint(0,10)
# # def random_num():
# #     attempts = 0
# #     while attempts < 5:
# #             user_num = input("Pick a number: ")
# #             attempts += 1
# #
# #             if user_num == number:
# #                 print("you win!!")
# #             else:
# #               print("Try again")
# #               if attempts == 5:
# #                 print(f"you loose the number was {number}")
# #
# #
# # random_num()
#
# path = Path()
# for file in path.glob('*'):
#     print(file)
wb = xl.load_workbook('')
sheet = wb['Sheet1']
cell = sheet['a1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)